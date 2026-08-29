#!/usr/bin/env python3
"""
Procesa una extracción TARIC de CIRCABC y la carga en Supabase.

Lo ejecuta GitHub Actions cuando el portal sube los ficheros al bucket
'taric-updates'. La carga va dentro de UNA transacción: si algo falla,
el catálogo antiguo sigue intacto.
"""
import os, re, sys, csv, json, subprocess, collections
from datetime import datetime, timezone
import requests
import openpyxl

def _env(nombre):
    v = os.environ.get(nombre, '').strip()
    if not v:
        print(f'[taric] FALTA la variable {nombre}. Revisa los secrets del repo '
              f'y que el campo id/extraccion del "Run workflow" no vaya vacío.', flush=True)
        sys.exit(2)
    return v

SB   = _env('SUPABASE_URL').rstrip('/')
SRV  = _env('SUPABASE_SERVICE_ROLE_KEY')
DB   = _env('SUPABASE_DB_URL')
ACT  = _env('ACT_ID')
EXTR = _env('EXTRACCION')

# Con la clave publishable, RLS se aplica y todo devuelve 200 con cero filas:
# ni error ni datos. Es el fallo más difícil de diagnosticar, así que se corta aquí.
if SRV.startswith('sb_publishable_') or SRV.startswith('sbp_'):
    print('[taric] ERROR: SUPABASE_SERVICE_ROLE_KEY contiene la clave PUBLISHABLE, '
          'no la service role. Con esa clave el Action no ve ni las filas ni los '
          'ficheros del bucket. Cópiala de Project Settings → API Keys → service_role '
          '(empieza por sb_secret_ o eyJ).', flush=True)
    sys.exit(2)

H = {'apikey': SRV, 'Authorization': f'Bearer {SRV}'}
WORK = 'taric_work'
os.makedirs(WORK, exist_ok=True)


def estado(**campos):
    """Actualiza la fila de control. Avisa si no llega a escribir nada:
    un id inexistente devolvería 200 sin tocar ninguna fila."""
    try:
        r = requests.patch(f'{SB}/rest/v1/taric_actualizaciones?id=eq.{ACT}',
                           headers={**H, 'Content-Type': 'application/json',
                                    'Prefer': 'return=representation'},
                           json=campos, timeout=30)
        if r.status_code >= 300:
            print(f'[taric] aviso: no se pudo escribir el estado ({r.status_code}) {r.text[:200]}', flush=True)
        elif r.json() == []:
            print(f'[taric] AVISO: no existe ninguna fila con id={ACT} en '
                  f'taric_actualizaciones. El proceso sigue, pero el portal no '
                  f'verá el progreso.', flush=True)
    except Exception as e:
        print(f'[taric] aviso: fallo al escribir estado: {e}', flush=True)


def log(msg):
    print(f'[taric] {msg}', flush=True)


def listar_bucket(prefijo=''):
    """Lista lo que hay realmente en el bucket, para diagnosticar rutas."""
    try:
        r = requests.post(f'{SB}/storage/v1/object/list/taric-updates',
                          headers={**H, 'Content-Type': 'application/json'},
                          json={'prefix': prefijo, 'limit': 100,
                                'sortBy': {'column': 'name', 'order': 'asc'}},
                          timeout=60)
        if r.status_code >= 300:
            return f'(no se pudo listar: {r.status_code} {r.text[:150]})'
        items = r.json()
        if not items:
            return '(vacío)'
        return ', '.join(i.get('name', '?') for i in items)
    except Exception as e:
        return f'(error al listar: {e})'


def descargar(nombre):
    """Baja un fichero del bucket privado. Devuelve None si no está."""
    url = f'{SB}/storage/v1/object/taric-updates/{EXTR}/{nombre}'
    destino = f'{WORK}/{nombre}'
    log(f'descargando {nombre}...')
    with requests.get(url, headers=H, stream=True, timeout=900) as r:
        if r.status_code in (400, 404):
            cuerpo = r.content[:200].decode('utf-8', 'replace')
            log(f'  no encontrado ({r.status_code}) {cuerpo}')
            return None
        r.raise_for_status()
        with open(destino, 'wb') as f:
            for chunk in r.iter_content(1 << 20):
                f.write(chunk)
    log(f'  {os.path.getsize(destino)/1e6:.1f} MB')
    return destino


# ══════════ Categorías paraduaneras ══════════
CAT = {
    '726': ('OZONO', 'Sustancias que agotan la capa de ozono'),
    '724': ('FGAS', 'Gases fluorados de efecto invernadero'),
    '711': ('DOBLE_USO', 'Productos y tecnologías de doble uso'),
    '738': ('MILITAR', 'Lista Común Militar (CML)'),
    '755': ('RESIDUOS', 'Traslado de residuos'),
    '734': ('PATRIMONIO', 'Bienes culturales / patrimonio'),
    '710': ('CITES', 'CITES - especies protegidas'),
    '410': ('VETERINARIO', 'Control veterinario (CHED-A/P)'),
    '415': ('FITOSANITARIO', 'Control oficial de vegetales (CHED-PP)'),
    '722': ('ALIMENTARIO', 'Restricción alimentos y piensos'),
    '750': ('ECOLOGICO', 'Control de productos ecológicos'),
    '761': ('REACH', 'REACH - sustancias químicas'),
    '769': ('POP', 'Contaminantes orgánicos persistentes'),
    '748': ('MERCURIO', 'Mercurio'),
    '713': ('OMG', 'Organismos modificados genéticamente'),
    '747': ('FLEGT', 'Madera - licencia FLEGT'),
    '770': ('FLEGT', 'Madera - licencia FLEGT (Ghana)'),
    '719': ('PESCA_INDNR', 'Pesca ilegal no declarada (INDNR)'),
    '746': ('FOCA', 'Productos derivados de la foca'),
    '745': ('PIEL_PERRO_GATO', 'Piel de perro y gato'),
    '728': ('LUJO', 'Bienes de lujo (sanciones)'),
    '705': ('TORTURA', 'Bienes para tortura o represión'),
    '712': ('EEI', 'Especies exóticas invasoras'),
    '775': ('CBAM', 'Mecanismo de ajuste en frontera por carbono'),
    '730': ('PRE_EXPORT', 'Verificación previa a la exportación'),
    '277': ('PROHIBICION', 'Prohibición de importación'),
    '465': ('RESTRICCION', 'Restricción de despacho a libre práctica'),
    '475': ('RESTRICCION', 'Restricción de despacho a libre práctica'),
    '474': ('LIMITE_CUANT', 'Limitación cuantitativa'),
    '552': ('ANTIDUMPING', 'Derecho antidumping definitivo'),
    '551': ('ANTIDUMPING', 'Derecho antidumping provisional'),
    '554': ('ANTISUBVENCION', 'Derecho compensatorio definitivo'),
    '564': ('ANTIDUMPING_REG', 'Registro antidumping/compensatorio'),
    '109': ('UNIDAD_SUPL', 'Unidad suplementaria'),
    '110': ('UNIDAD_SUPL', 'Unidad suplementaria (importación)'),
}
SANCIONES = {'707', '714', '731', '732', '760', '762', '763'}


def categoria_de(codigo, base):
    if codigo in CAT:
        return CAT[codigo]
    if codigo in SANCIONES:
        b = base or ''
        if any(x in b for x in ('0833/14', '0263/22', '0692/14', '0813/25')):
            return ('SANCION_RU', 'Sanciones Rusia / territorios ocupados de Ucrania')
        if '0267/12' in b:  return ('SANCION_IR', 'Sanciones Irán')
        if '1509/17' in b:  return ('SANCION_KP', 'Sanciones Corea del Norte')
        if '0765/06' in b or '1793/19' in b: return ('SANCION_BY', 'Sanciones Bielorrusia')
        return ('SANCION', 'Sanciones / medidas restrictivas')
    return None


def procesar_nomenclatura(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True); next(it)
    filas = []
    for r in it:
        m = re.match(r'^(\d{10})\s+(\d{2})$', str(r[0] or '').strip())
        if not m:
            continue
        cod, suf = m.group(1), m.group(2)
        filas.append([cod, suf, cod[:8], cod[:4], cod[:2],
                      1 if suf == '80' else 0, r[4], (r[5] or '').strip(),
                      (r[6] or '').strip(), r[1], r[2]])
    wb.close()
    ruta_csv = f'{WORK}/nom.csv'
    with open(ruta_csv, 'w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerows(filas)
    log(f'nomenclatura: {len(filas):,} filas')
    return ruta_csv, len(filas)


def procesar_medidas(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True)
    vistos, filas, n = set(), [], 0
    for sh in wb:
        it = sh.iter_rows(values_only=True); next(it, None)
        for r in it:
            n += 1
            # Latido cada 200.000 filas: sin esto el proceso pasa 10 minutos
            # mudo y no hay forma de saber si avanza o está colgado.
            if n % 200000 == 0:
                log(f'  ...{n:,} filas leídas, {len(filas):,} retenidas')
                estado(mensaje=f'Procesando medidas: {n:,} filas leídas')
            tipo = str(r[11]).strip() if r[11] else ''
            c = categoria_de(tipo, r[8])
            if not c:
                continue
            cod = str(r[0] or '').strip()
            if not re.match(r'^\d{10}$', cod):
                continue
            origen = (r[6] or '').strip()
            k = (cod, c[0], origen)
            if k in vistos:
                continue
            vistos.add(k)
            derecho = (r[9] or '').strip()
            certs = sorted({m.group(1) + m.group(2)
                            for m in re.finditer(r'\b([A-Z])-?(\d{3})\b', derecho)})
            reales = [x for x in certs if x[0] in 'NCLAU']
            erga = 1 if origen.upper() in ('ERGA OMNES', '', 'ALL THIRD COUNTRIES') else 0
            if reales:                              exig = 'DOCUMENTO'
            elif certs:                             exig = 'ATESTACION_Y'
            elif c[0] in ('PROHIBICION', 'RESTRICCION', 'LIMITE_CUANT'): exig = 'PROHIBICION'
            elif c[0] in ('ANTIDUMPING', 'ANTISUBVENCION', 'UNIDAD_SUPL'): exig = 'ARANCELARIA'
            else:                                   exig = 'REVISAR'
            filas.append([cod, cod[:8], c[0], c[1], tipo, (r[7] or '').strip(),
                          origen, (r[10] or '').strip(), (r[8] or '').strip(), derecho,
                          r[3].date().isoformat() if hasattr(r[3], 'date') else None,
                          r[4].date().isoformat() if hasattr(r[4], 'date') else None,
                          '|'.join(certs), erga, exig])
    wb.close()
    ruta_csv = f'{WORK}/med.csv'
    with open(ruta_csv, 'w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerows(filas)
    log(f'medidas: {n:,} leídas, {len(filas):,} retenidas')
    return ruta_csv, len(filas)


def procesar_casilla44(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True)
    it = wb[wb.sheetnames[0]].iter_rows(values_only=True); next(it)
    filas = [[str(r[0]).strip(), (r[2] or '').strip(), r[3], r[5]]
             for r in it if r[1] == 'ES' and not r[5]]
    wb.close()
    ruta_csv = f'{WORK}/c44.csv'
    with open(ruta_csv, 'w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerows(filas)
    log(f'casilla 44: {len(filas):,} códigos')
    return ruta_csv, len(filas)


def cargar(nom_csv, med_csv, c44_csv):
    """Todo en una transacción: o entra completo o no entra nada."""
    sql = ['\\set ON_ERROR_STOP on', 'BEGIN;']
    sql.append('DELETE FROM public.taric_nomenclatura;')
    sql.append(f"\\copy public.taric_nomenclatura (codigo,sufijo,nc8,partida4,capitulo,declarable,nivel,indent,descripcion,inicio,fin) FROM '{nom_csv}' WITH (FORMAT csv)")
    sql.append('DELETE FROM public.taric_medidas;')
    sql.append(f"\\copy public.taric_medidas (codigo,nc8,categoria,etiqueta,tipo_medida,desc_medida,origen,origen_code,base_legal,derecho,inicio,fin,certificados,erga_omnes,exigencia) FROM '{med_csv}' WITH (FORMAT csv)")
    if c44_csv:
        sql.append('DELETE FROM public.taric_casilla44;')
        sql.append(f"\\copy public.taric_casilla44 (codigo,descripcion,inicio,fin) FROM '{c44_csv}' WITH (FORMAT csv)")
    sql.append('COMMIT;')

    ruta = f'{WORK}/carga.sql'
    open(ruta, 'w', encoding='utf-8').write('\n'.join(sql) + '\n')
    log('cargando en Supabase...')
    r = subprocess.run(['psql', DB, '-v', 'ON_ERROR_STOP=1', '-f', ruta],
                       capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f'psql falló: {r.stderr[-1500:]}')
    log('carga completada')


def main():
    log(f'parámetros · id={ACT} · extracción={EXTR} · proyecto={SB}')
    estado(estado='procesando', mensaje='Comprobando conexión')

    # Fallar rápido: mejor un error a los 2 segundos que tras 10 minutos
    # de parseo. Verifica de paso que el pooler y la contraseña son buenos.
    r = subprocess.run(['psql', DB, '-tAc', 'SELECT count(*) FROM public.taric_medidas;'],
                       capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f'No se pudo conectar a la base: {r.stderr.strip()[-400:]}')
    log(f'conexión OK · medidas actuales: {r.stdout.strip()}')

    estado(mensaje='Descargando ficheros del bucket')

    nom = descargar('Nomenclature_EN.xlsx') or descargar('Nomenclature_ES.xlsx')
    med = descargar('TARIC_measures.xlsx')
    c44 = descargar('Box_44_codes_of_the_SAD.xlsx')
    if not nom or not med:
        log(f'contenido de taric-updates/{EXTR}/ : {listar_bucket(EXTR + "/")}')
        log(f'carpetas en la raíz del bucket   : {listar_bucket("")}')
        raise RuntimeError(
            f'Faltan Nomenclature_*.xlsx o TARIC_measures.xlsx en taric-updates/{EXTR}/. '
            f'Revisa en el log de arriba qué hay realmente en el bucket.')

    estado(mensaje='Procesando nomenclatura')
    nom_csv, n_nom = procesar_nomenclatura(nom)

    estado(mensaje='Procesando medidas (puede tardar varios minutos)')
    med_csv, n_med = procesar_medidas(med)

    n_c44, c44_csv = 0, None
    if c44:
        c44_csv, n_c44 = procesar_casilla44(c44)

    if n_nom < 10000 or n_med < 10000:
        raise RuntimeError(f'Cifras sospechosamente bajas ({n_nom} / {n_med}). No se carga.')

    estado(mensaje='Cargando en la base de datos')
    cargar(os.path.abspath(nom_csv), os.path.abspath(med_csv),
           os.path.abspath(c44_csv) if c44_csv else None)

    estado(estado='ok', filas_nom=n_nom, filas_med=n_med, filas_c44=n_c44,
           mensaje=f'Catálogo {EXTR} cargado correctamente',
           terminado_en=datetime.now(timezone.utc).isoformat())
    log('LISTO')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        log(f'ERROR: {e}')
        estado(estado='error', mensaje=str(e)[:500],
               terminado_en=datetime.now(timezone.utc).isoformat())
        sys.exit(1)
